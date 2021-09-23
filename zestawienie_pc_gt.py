import os, re, datetime, time, email, extract_msg
from openpyxl import Workbook
from email import policy
from email.parser import BytesParser


wb = Workbook()
ws = wb.active

dirMain = 'C:/Users/swidr/PV Poland Sp. z o.o/Techniczny - Pompy Ciepła/Oferty/'
folders = os.listdir(dirMain)

citiesList = []
nameList = []
dateList = []
mailList = []

cityRegex = re.compile('([A-ZŚĄĘÓŃĆŹŻŁ]{2,})')
nameRegex = re.compile('[A-ZŚĄĘÓŃŹĆŻŁ]+[a-ząśęłóćńźż]+')
mailRegex = re.compile('[a-z]+\.[a-z]+@pvpoland.pl')

for dirName in folders:
    citiesList.append(' '.join(cityRegex.findall(dirName)))
    nameList.append(' '.join(nameRegex.findall(dirName)))
    for plik in os.listdir('C:/Users/swidr/PV Poland Sp. z o.o/Techniczny - Pompy Ciepła/Oferty/{}'.format(dirName)):
#        if ('.eml') in plik:
#            with open('C:/Users/swidr/Oferty/{}/{}'.format(dirName, plik), 'rb') as fp:
#                msg = BytesParser(policy=policy.default).parse(fp)
#                mailList.append(msg['from'])
#        elif ('.msg') in plik:
#            f = 'C:/Users/swidr/Oferty/{}/{}'.format(dirName, plik)
#            msg = extract_msg.Message(f)
#            mailList.append(msg.sender)
        if ('Oferta' and '.pdf') in plik:
            dateList.append(datetime.datetime.strptime(time.ctime(os.path.getctime('C:/Users/swidr/PV Poland Sp. z o.o/Techniczny - Pompy Ciepła/Oferty/{}/{}'.format(dirName, plik))), '%a %b %d %H:%M:%S %Y').strftime('%d.%m.%Y'))
        else:
            mailList.append(' ')         
    #for filename in os.listdir('C:/Users/swidr/Oferty/{}'.format(dirName)):
        #if filename.startswith("Oferta") and filename.endswith(".pdf"):
            #dateList.append(datetime.datetime.strptime(time.ctime(os.path.getctime('C:/Users/swidr/Oferty/{}/{}'.format(dirName, filename))), '%a %b %d %H:%M:%S %Y').strftime('%d.%m.%Y'))
        #if filename.endswith(".eml"):
            #with open('C:/Users/swidr/Oferty/{}/{}'.format(dirName, filename), 'rb') as fp:
                #msg = BytesParser(policy=policy.default).parse(fp)
                #mailList.append(msg['from'])
                #mailText = msg.get_body(preferencelist=('plain')).get_content()
                #mailList.append(mailRegex.match(mailText))

i = 0

for index, data in enumerate(citiesList):
    ws.cell(index + 2, 1).value = citiesList[i]
    ws.cell(index + 2, 2).value = nameList[i]
    ws.cell(index + 2, 3).value = dateList[i]
    ws.cell(index + 2, 4).value = mailList[i]
    i = i + 1
    

wb.save('Zestawienie.xlsx')

